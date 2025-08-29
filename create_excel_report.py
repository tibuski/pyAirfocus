import json
import pandas as pd
from collections import defaultdict
import os
import argparse

def load_json_data(file_path):
    """Load JSON data from file"""
    with open(file_path, 'r', encoding='utf-8') as file:
        return json.load(file)

def extract_workspace_data(data):
    """Extract workspace data and create a structured dataset"""
    workspaces = []
    
    # Skip _GLOBAL_TOTALS and process cluster data
    for cluster_key, cluster_data in data.items():
        if cluster_key == "_GLOBAL_TOTALS":
            continue
            
        # Handle nested structure (cluster > workspace)
        for workspace_key, workspace_data in cluster_data.items():
            workspace_info = {
                'Cluster': cluster_key,
                'Workspace': workspace_key,
                'Full Path': f"{cluster_key} > {workspace_key}"
            }
            
            # Extract workspace totals if available
            if '_WORKSPACE_TOTALS' in workspace_data:
                totals = workspace_data['_WORKSPACE_TOTALS']
                workspace_info.update({
                    'Total Filled Fields': totals.get('total_true_fields', 0),
                    'Total Unset Fields': totals.get('total_false_fields', 0),
                    'Total Items': totals.get('total_items', 0),
                    'Total Unique Fields': totals.get('total_unique_fields', 0)
                })
            else:
                # Initialize with zeros if no totals found
                workspace_info.update({
                    'Total Filled Fields': 0,
                    'Total Unset Fields': 0,
                    'Total Items': 0,
                    'Total Unique Fields': 0
                })
            
            # Extract field statistics and status counts
            field_stats = defaultdict(int)
            status_counts = defaultdict(int)
            all_fields = set()
            
            # Process individual items in the workspace
            for item_key, item_data in workspace_data.items():
                if item_key == '_WORKSPACE_TOTALS':
                    continue
                
                # Count status occurrences
                if '_status' in item_data:
                    status_counts[item_data['_status']] += 1
                
                # Count field true/false occurrences
                for field_key, field_value in item_data.items():
                    if field_key != '_status' and isinstance(field_value, bool):
                        all_fields.add(field_key)
                        if field_value:
                            field_stats[f"{field_key} True"] += 1
                        else:
                            field_stats[f"{field_key} False"] += 1
            
            # Add field statistics to workspace info
            workspace_info.update(field_stats)
            
            # Add status counts to workspace info
            for status, count in status_counts.items():
                workspace_info[f"Status {status.replace('_', ' ')}"] = count
            
            workspaces.append(workspace_info)
    
    return workspaces

def create_excel_report(workspaces_data, output_file):
    """Create Excel report with multiple sheets"""
    
    # Create main dataframe
    df = pd.DataFrame(workspaces_data)
    
    # Fill NaN values with 0 for numeric columns
    numeric_columns = df.select_dtypes(include=['number']).columns
    df[numeric_columns] = df[numeric_columns].fillna(0)
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Summary totals only (moved to first position)
        summary_cols = ['Cluster', 'Workspace', 'Full Path', 'Total Filled Fields', 
                       'Total Unset Fields', 'Total Items', 'Total Unique Fields']
        df_summary = df[summary_cols].copy()
        df_summary.to_excel(writer, sheet_name='Summary Totals', index=False)
        
        # Sheet 2: Status counts
        status_cols = ['Cluster', 'Workspace', 'Full Path'] + [col for col in df.columns if col.startswith('Status ')]
        if len(status_cols) > 3:  # Only create if status columns exist
            df_status = df[status_cols].copy()
            df_status.to_excel(writer, sheet_name='Status Counts', index=False)
        
        # Sheet 3: Field statistics (True/False counts)
        field_cols = ['Cluster', 'Workspace', 'Full Path'] + [col for col in df.columns if col.endswith(' True') or col.endswith(' False')]
        if len(field_cols) > 3:  # Only create if field columns exist
            df_fields = df[field_cols].copy()
            df_fields.to_excel(writer, sheet_name='Field Statistics', index=False)
        
        # Sheet 4: Complete data (moved to last position)
        df.to_excel(writer, sheet_name='Complete Data', index=False)
        
        # Add conditional formatting to Summary Totals sheet
        summary_ws = writer.sheets['Summary Totals']
        
        # Find column indices for conditional formatting
        summary_header = list(df_summary.columns)
        filled_col_idx = summary_header.index('Total Filled Fields') + 1  # +1 for Excel 1-based indexing
        unset_col_idx = summary_header.index('Total Unset Fields') + 1
        items_col_idx = summary_header.index('Total Items') + 1
        unique_fields_col_idx = summary_header.index('Total Unique Fields') + 1
        
        # Convert column indices to letters
        from openpyxl.utils import get_column_letter
        filled_col_letter = get_column_letter(filled_col_idx)
        unset_col_letter = get_column_letter(unset_col_idx)
        items_col_letter = get_column_letter(items_col_idx)
        unique_fields_col_letter = get_column_letter(unique_fields_col_idx)
        
        # Define the data range (excluding header)
        data_start_row = 2
        data_end_row = len(df_summary) + 1
        
        # Import required formatting classes
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        
        # Define fill colors for conditional formatting - 5 red to 5 green evolution
        # Red variations (poor performance) - 5 levels
        dark_red_fill = PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid')      # Dark red
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')           # Red
        medium_red_fill = PatternFill(start_color='FFD9D9', end_color='FFD9D9', fill_type='solid')    # Medium red
        light_red_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')     # Light red
        very_light_red_fill = PatternFill(start_color='FFF2F2', end_color='FFF2F2', fill_type='solid') # Very light red
        
        # Green variations (excellent performance) - 5 levels
        very_light_green_fill = PatternFill(start_color='F2FFF2', end_color='F2FFF2', fill_type='solid') # Very light green
        light_green_fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')    # Light green
        medium_green_fill = PatternFill(start_color='D9E6D9', end_color='D9E6D9', fill_type='solid')   # Medium green
        green_fill = PatternFill(start_color='CCE6CC', end_color='CCE6CC', fill_type='solid')           # Green
        dark_green_fill = PatternFill(start_color='B3D9B3', end_color='B3D9B3', fill_type='solid')     # Dark green
        
        # Apply conditional formatting row by row for Total Filled Fields
        for row in range(data_start_row, data_end_row + 1):
            # Calculate percentage thresholds for this row
            # Formula: =IF(AND(D2>0,E2>0),D2/(D2*E2),0) - but we'll use the actual values
            
            # Get the actual values for this row
            items_value = df_summary.iloc[row - data_start_row]['Total Items']
            unique_fields_value = df_summary.iloc[row - data_start_row]['Total Unique Fields']
            
            if items_value > 0 and unique_fields_value > 0:
                max_possible_fields = items_value * unique_fields_value
                
                # Apply conditional formatting based on percentage thresholds with 5 red to 5 green evolution
                # 0-10% dark red, 11-20% red, 21-30% medium red, 31-40% light red, 41-50% very light red,
                # 51-60% very light green, 61-70% light green, 71-80% medium green, 81-90% green, 91-100% dark green
                
                # Dark red: 0-10%
                dark_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'0', f'{max_possible_fields * 0.1}'],
                    fill=dark_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    dark_red_rule
                )
                
                # Red: 11-20%
                red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.1 + 1}', f'{max_possible_fields * 0.2}'],
                    fill=red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    red_rule
                )
                
                # Medium red: 21-30%
                medium_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.2 + 1}', f'{max_possible_fields * 0.3}'],
                    fill=medium_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    medium_red_rule
                )
                
                # Light red: 31-40%
                light_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.3 + 1}', f'{max_possible_fields * 0.4}'],
                    fill=light_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    light_red_rule
                )
                
                # Very light red: 41-50%
                very_light_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.4 + 1}', f'{max_possible_fields * 0.5}'],
                    fill=very_light_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    very_light_red_rule
                )
                
                # Very light green: 51-60%
                very_light_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.5 + 1}', f'{max_possible_fields * 0.6}'],
                    fill=very_light_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    very_light_green_rule
                )
                
                # Light green: 61-70%
                light_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.6 + 1}', f'{max_possible_fields * 0.7}'],
                    fill=light_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    light_green_rule
                )
                
                # Medium green: 71-80%
                medium_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.7 + 1}', f'{max_possible_fields * 0.8}'],
                    fill=medium_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    medium_green_rule
                )
                
                # Green: 81-90%
                green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.8 + 1}', f'{max_possible_fields * 0.9}'],
                    fill=green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    green_rule
                )
                
                # Dark green: 91-100%
                dark_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.9 + 1}', f'{max_possible_fields}'],
                    fill=dark_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{filled_col_letter}{row}',
                    dark_green_rule
                )
        
        # Apply conditional formatting row by row for Total Unset Fields (inverted evolution)
        for row in range(data_start_row, data_end_row + 1):
            # Get the actual values for this row
            items_value = df_summary.iloc[row - data_start_row]['Total Items']
            unique_fields_value = df_summary.iloc[row - data_start_row]['Total Unique Fields']
            
            if items_value > 0 and unique_fields_value > 0:
                max_possible_fields = items_value * unique_fields_value
                
                # Apply conditional formatting based on percentage thresholds (inverted evolution)
                # 0-10% dark green, 11-20% green, 21-30% medium green, 31-40% light green, 41-50% very light green,
                # 51-60% very light red, 61-70% light red, 71-80% medium red, 81-90% red, 91-100% dark red
                
                # Dark green: 0-10%
                dark_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'0', f'{max_possible_fields * 0.1}'],
                    fill=dark_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    dark_green_rule
                )
                
                # Green: 11-20%
                green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.1 + 1}', f'{max_possible_fields * 0.2}'],
                    fill=green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    green_rule
                )
                
                # Medium green: 21-30%
                medium_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.2 + 1}', f'{max_possible_fields * 0.3}'],
                    fill=medium_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    medium_green_rule
                )
                
                # Light green: 31-40%
                light_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.3 + 1}', f'{max_possible_fields * 0.4}'],
                    fill=light_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    light_green_rule
                )
                
                # Very light green: 41-50%
                very_light_green_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.4 + 1}', f'{max_possible_fields * 0.5}'],
                    fill=very_light_green_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    very_light_green_rule
                )
                
                # Very light red: 51-60%
                very_light_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.5 + 1}', f'{max_possible_fields * 0.6}'],
                    fill=very_light_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    very_light_red_rule
                )
                
                # Light red: 61-70%
                light_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.6 + 1}', f'{max_possible_fields * 0.7}'],
                    fill=light_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    light_red_rule
                )
                
                # Medium red: 71-80%
                medium_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.7 + 1}', f'{max_possible_fields * 0.8}'],
                    fill=medium_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    medium_red_rule
                )
                
                # Red: 81-90%
                red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.8 + 1}', f'{max_possible_fields * 0.9}'],
                    fill=red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    red_rule
                )
                
                # Dark red: 91-100%
                dark_red_rule = CellIsRule(
                    operator='between',
                    formula=[f'{max_possible_fields * 0.9 + 1}', f'{max_possible_fields}'],
                    fill=dark_red_fill
                )
                summary_ws.conditional_formatting.add(
                    f'{unset_col_letter}{row}',
                    dark_red_rule
                )
        
        # Apply modern orange theme to all sheets
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Define orange-based theme colors
        header_fill = PatternFill(start_color='FF8C42', end_color='FF8C42', fill_type='solid')  # Modern orange
        alt_row_fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')  # Light orange
        header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')  # White text
        data_font = Font(name='Segoe UI', size=10, color='333333')  # Dark gray text
        border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        alignment = Alignment(horizontal='left', vertical='center')
        
        # Apply theme to all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Style header row
            for cell in worksheet[1]:  # First row (headers)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
                cell.border = border
            
            # Style data rows with alternating colors
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                fill_color = alt_row_fill if row_idx % 2 == 0 else PatternFill()  # Alternate rows
                for cell in row:
                    cell.font = data_font
                    cell.alignment = alignment
                    cell.border = border
                    if row_idx % 2 == 0:
                        cell.fill = fill_color
        
        # Auto-adjust column widths for all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function to process JSON and create Excel report"""
    
    # Set up command line argument parsing
    parser = argparse.ArgumentParser(
        description='Process JSON data and create Excel report with workspace analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python create_excel_report.py field_statuses.json report.xlsx
  python create_excel_report.py data.json analysis.xlsx
        """
    )
    
    parser.add_argument('json_file', 
                       help='Path to the input JSON file containing workspace data')
    parser.add_argument('excel_file', 
                       help='Path for the output Excel report file')
    parser.add_argument('-v', '--verbose', 
                       action='store_true',
                       help='Enable verbose output')
    
    args = parser.parse_args()
    
    json_file = args.json_file
    excel_file = args.excel_file
    
    # Check if JSON file exists
    if not os.path.exists(json_file):
        print(f"Error: {json_file} not found")
        return
    
    try:
        # Load and process data
        print("Loading JSON data...")
        data = load_json_data(json_file)
        
        print("Extracting workspace data...")
        workspaces = extract_workspace_data(data)
        
        print(f"Found {len(workspaces)} workspaces")
        
        # Create Excel report
        print("Creating Excel report...")
        create_excel_report(workspaces, excel_file)
        
        print(f"Excel report created successfully: {excel_file}")
        print("\nReport contains the following sheets:")
        print("1. Summary Totals - Workspace totals only")
        print("2. Status Counts - Item status distribution")
        print("3. Field Statistics - Field true/false counts")
        print("4. Complete Data - All extracted data")
        
        # Display summary statistics
        df = pd.DataFrame(workspaces)
        print(f"\nSummary Statistics:")
        print(f"Total workspaces: {len(workspaces)}")
        print(f"Total items across all workspaces: {df['Total Items'].sum()}")
        print(f"Total filled fields: {df['Total Filled Fields'].sum()}")
        print(f"Total unset fields: {df['Total Unset Fields'].sum()}")
        
    except Exception as e:
        print(f"Error processing data: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
